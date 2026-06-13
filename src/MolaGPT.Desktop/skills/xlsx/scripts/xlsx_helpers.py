"""xlsx_helpers.py — 可复用的 Excel 构建/读取（MolaGPT 内置技能）

用法：
    import sys; sys.path.append(r"<技能目录>/scripts")
    from xlsx_helpers import WorkbookBuilder, read_xlsx
    wb = WorkbookBuilder("report.xlsx")
    wb.sheet("销售")
    wb.table(["月份","销量","单价"], [["1月",120,9.9],["2月",150,9.9]],
             money_cols=[2], formula_total_col=None)
    wb.bar_chart(title="月度销量", cat_col=0, val_col=1, n_rows=2)
    wb.close()
"""

import openpyxl


class WorkbookBuilder:
    """基于 xlsxwriter：支持表头样式、数字格式、公式、图表、自动列宽。"""

    def __init__(self, path):
        import xlsxwriter
        self.path = path
        self.wb = xlsxwriter.Workbook(path)
        self.ws = None
        self._fmt = {
            "header": self.wb.add_format({"bold": True, "bg_color": "#1E2761", "font_color": "white", "border": 1, "align": "center"}),
            "cell": self.wb.add_format({"border": 1}),
            "money": self.wb.add_format({"num_format": "￥#,##0.00", "border": 1}),
            "int": self.wb.add_format({"num_format": "#,##0", "border": 1}),
            "pct": self.wb.add_format({"num_format": "0.0%", "border": 1}),
        }
        self._last = {}

    def sheet(self, name):
        self.ws = self.wb.add_worksheet(name)
        self._last["name"] = name
        return self.ws

    def table(self, header, rows, money_cols=(), int_cols=(), pct_cols=(), start_row=0):
        money_cols, int_cols, pct_cols = set(money_cols), set(int_cols), set(pct_cols)
        for c, h in enumerate(header):
            self.ws.write(start_row, c, str(h), self._fmt["header"])
        widths = [len(str(h)) + 2 for h in header]
        for r, row in enumerate(rows, start=start_row + 1):
            for c, v in enumerate(row):
                fmt = (self._fmt["money"] if c in money_cols else
                       self._fmt["int"] if c in int_cols else
                       self._fmt["pct"] if c in pct_cols else self._fmt["cell"])
                if isinstance(v, (int, float)):
                    self.ws.write_number(r, c, v, fmt)
                else:
                    self.ws.write(r, c, str(v), fmt)
                widths[c] = max(widths[c], len(str(v)) + 2)
        for c, w in enumerate(widths):
            self.ws.set_column(c, c, min(w, 40))
        self._last.update(header=header, nrows=len(rows), start_row=start_row)
        return self

    def bar_chart(self, title, cat_col, val_col, n_rows, anchor="H2"):
        chart = self.wb.add_chart({"type": "column"})
        name = self._last.get("name", "Sheet1")
        first = self._last.get("start_row", 0) + 2  # 1-based, skip header
        last = first + n_rows - 1
        col_letter = chr(ord("A") + cat_col)
        val_letter = chr(ord("A") + val_col)
        chart.add_series({
            "categories": f"='{name}'!${col_letter}${first}:${col_letter}${last}",
            "values":     f"='{name}'!${val_letter}${first}:${val_letter}${last}",
            "name": title,
        })
        chart.set_title({"name": title})
        self.ws.insert_chart(anchor, chart)
        return self

    def close(self):
        self.wb.close()
        return self.path


def read_xlsx(path, values_only=True):
    """读取所有工作表为 {名称: 二维列表}。"""
    wb = openpyxl.load_workbook(path, data_only=values_only)
    out = {}
    for ws in wb.worksheets:
        out[ws.title] = [list(row) for row in ws.iter_rows(values_only=True)]
    return out
